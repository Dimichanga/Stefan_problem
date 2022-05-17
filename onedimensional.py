from classes import time_, geometry, Boundary 
import openpyxl

def onedimensional_Stefan (str1, str2, str3, onedimens, Left, times):

    with open (str1,'r') as f:
        initcond = float(f.read())
    with open (str2, "r") as f:
        coef_therm_cond = float(f.read())
    with open (str3, "r") as f:
        enthalpy = float(f.read())

    numb = int(input("Enter the number of steps in the space(by default enter 10): "))
    tau = float(input("Enter time step for the calculation (by default enter 0.5): "))
    h = (onedimens.length / numb) 
    eps = float(input("Enter epsilon for the accuracy of iterations of time steps: "))
    phase_index = int(input("Enter the coordinate of the phase transition boundary: "))

    Nl=len(Left.table)
    u=[None] * (numb+1)
    u_=[None] * (numb+1)
    alpha = [None] * (numb+1)
    beta = [None] * (numb+1)
    gamma = [None] * (numb+1)
        
    times.duration = Left.table[Nl-1][0]
    times.set_step(tau)
    times.set_prev_step(0)
    
    u[0] = Left.table[0][1]
    for i in range(1,numb+1):
            if i < phase_index: 
                u[i] = initcond
            elif i == phase_index:
                u[i] = 0
            else:
                u[i]=0
    
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row = 1, column = 1).value = 't/x' 
    for i in range (numb+1):
        sheet.cell(row = 1, column = i+2).value = h * i
     
    count = 0
    while times.actual < times.duration:

        sheet.cell(row = count + 2, column = 1).value = times.actual
        for i in range (phase_index+1):
            sheet.cell(row = count+2, column = i+2).value = u[i]
        count += 1
        
        for i in range (numb+1):
            u_[i] = u[i]

        if (phase_index < numb): 
            phase_index += 1
        
        while (abs(times.step - times.prev_step) >= eps):
            A = (- tau) / (h**2)
            B = 1 + (2 * tau/(h**2))
            C = A
            u[0] = Left.get(times.actual)
            u[phase_index] =0
            gamma [1] = B
            alpha[1] = -C / gamma[1]
            beta[1] = (u_[1] - A*u[0]) / gamma[1]

            for i in range (2,phase_index):
                gamma[i] = B + A * alpha[i-1]
                if (i != phase_index-1):
                    beta[i] = (u_[i] - A*beta[i-1]) / gamma[i]
                    alpha[i] = -C / gamma[i]
                else:
                    beta[i] = (u_[i] - C*u[phase_index] - A*beta[i-1]) / gamma[i]
                    alpha[i] = 0
        
            u[phase_index-1] = beta[phase_index-1]
            for i in range (phase_index-2,0,-1):
                u[i] = alpha[i] * u[i+1] + beta[i]

            times.prev_step = times.step
            times.step = -((enthalpy / coef_therm_cond )*h + 0.5*(u[phase_index]-u_[phase_index])) / ((u[phase_index] - u[phase_index-1]) / h)

            

        times.actual += times.step

    sheet.cell(row = count + 3, column = 1).value = times.actual
    for i in range(numb+1):
        sheet.cell(row = count+3, column = i+2).value = u[i]

    book.save("Solution of the Stefan problem.xlsx")
    book.close()
